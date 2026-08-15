# make_sample.py —— 生成一个示例 Excel 文件
# 作用：让你不用手动建表，直接有个现成的 sales.xlsx 可以读
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "销售表"

# 写表头
ws["A1"] = "月份"
ws["B1"] = "销售额"
ws["C1"] = "利润"

# 写数据
data = [
    ("1月", 1000, 200),
    ("2月", 1500, 350),
    ("3月", 1200, 280),
    ("4月", 1800, 450),
]

for i, (month, sales, profit) in enumerate(data, start=2):
    ws.cell(row=i, column=1, value=month)
    ws.cell(row=i, column=2, value=sales)
    ws.cell(row=i, column=3, value=profit)

wb.save("sales.xlsx")
print("示例文件 sales.xlsx 已生成")
