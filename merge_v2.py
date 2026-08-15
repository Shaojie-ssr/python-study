# merge_v2.py —— 合并 + 按人统计，一次出两张 sheet
# 这个文件是"必做练习"完成后的扩展，让你感受"在原版上叠几行代码就能多出功能"
import os
from openpyxl import load_workbook, Workbook
from collections import defaultdict

# 1. 找文件
files = [f for f in os.listdir(".") if f.startswith("sales_") and f.endswith(".xlsx")]
files.sort()
print(f"找到 {len(files)} 个文件")

# 2. 创建工作簿
dst = Workbook()

# 表1：明细汇总
ws1 = dst.active
ws1.title = "明细"
ws1.append(["月份", "销售员", "销售额"])

# 字典暂存"按人汇总"
person_totals = defaultdict(int)

# 3. 逐个文件读 → 写
for f in files:
    month = f.replace("sales_", "").replace(".xlsx", "")
    src = load_workbook(f)
    src_ws = src.active
    for row in src_ws.iter_rows(min_row=2, values_only=True):
        name, amount = row
        ws1.append([month, name, amount])
        person_totals[name] += amount      # 顺便累加

# 4. 表2：按人汇总
ws2 = dst.create_sheet("按人汇总")
ws2.append(["销售员", "总销售额"])
for name, total in sorted(person_totals.items()):
    ws2.append([name, total])

# 5. 保存
dst.save("汇总v2.xlsx")
print("汇总v2.xlsx 已生成 ✅\n")

# 6. 看一眼第二个 sheet
print("--- 按人汇总 ---")
for row in ws2.iter_rows(values_only=True):
    print(row)