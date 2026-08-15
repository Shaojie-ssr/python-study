# merge_sheets.py —— 杀招：自动找所有 sales_*.xlsx，合并成一张大表
import os
from openpyxl import load_workbook, Workbook

# 1. 找当前目录下所有"销售"相关 xlsx
files = [f for f in os.listdir(".") if f.startswith("sales_") and f.endswith(".xlsx")]
files.sort()  # 按文件名排序，结果可预测

print(f"找到 {len(files)} 个文件：")
for f in files:
    print(f"  - {f}")
print()

# 2. 创建新工作簿 + 写汇总表头
dst = Workbook()
ws = dst.active
ws.title = "汇总"
ws.append(["月份", "销售员", "销售额"])

# 3. 逐个文件读 → 写
for f in files:
    month = f.replace("sales_", "").replace(".xlsx", "")  # 从文件名提取月份
    src = load_workbook(f)
    src_ws = src.active
    for row in src_ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        name, amount = row
        ws.append([month, name, amount])

# 4. 保存
dst.save("汇总.xlsx")
print("汇总.xlsx 已生成 ✅\n")

# 5. 自检：读回来看一眼
print("--- 汇总表内容 ---")
check = load_workbook("汇总.xlsx")
for row in check.active.iter_rows(values_only=True):
    print(row)