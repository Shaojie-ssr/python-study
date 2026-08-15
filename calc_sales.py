from openpyxl import load_workbook

wb = load_workbook("sales.xlsx")
ws = wb.active

total_sales = 0
total_profit = 0

for row in ws.iter_rows(min_row=2, values_only=True):  # 从第二行起，跳过表头
    month, sales, profit = row
    total_sales += sales
    total_profit += profit

print(f"总销售额：{total_sales}")
print(f"总利润：{total_profit}")
print(f"利润率：{total_profit / total_sales * 100:.1f}%")
