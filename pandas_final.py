"""pandas 综合实战：销售分析报告生成器
输入：monthly_sales/ 下的 6 个月销售表
输出：report/ 文件夹下
  - 销售分析报告_YYYY-MM-DD.xlsx（数据 + 文字总结 + 4 张图嵌入）
  - chart_trend.png / chart_person.png / chart_pie.png / chart_compare.png
"""
import matplotlib
matplotlib.use("Agg")
matplotlib.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei"]
matplotlib.rcParams["axes.unicode_minus"] = False

import matplotlib.pyplot as plt
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

# ===== 1. 准备 =====
report_dir = Path("report")
report_dir.mkdir(exist_ok=True)
date_str = datetime.now().strftime("%Y-%m-%d")
COLORS = ["#D85A30", "#EF9F27", "#378ADD", "#1D9E75", "#888780"]

# ===== 2. 读 6 个月数据 =====
files = sorted(Path("monthly_sales").glob("sales_*.xlsx"))
df = pd.concat([pd.read_excel(f) for f in files], ignore_index=True)
print(f"载入 {len(files)} 个月共 {len(df)} 条销售记录\n")

# ===== 3. 算核心指标 =====
total = df["销售额"].sum()
avg_per_deal = df["销售额"].mean()
n_deals = len(df)
by_person = df.groupby("销售员")["销售额"].sum().sort_values(ascending=False)
by_month = df.groupby("月份")["销售额"].sum()
top, top_amount = by_person.idxmax(), by_person.max()
top_pct = top_amount / total * 100
growth = (by_month.iloc[-1] / by_month.iloc[0] - 1) * 100

# ===== 4. 出 4 张图 =====
# 4.1 月度趋势
plt.figure(figsize=(10, 5))
plt.plot(by_month.index, by_month.values, marker="o", linewidth=2.5,
         color="#378ADD", markersize=8)
plt.title("月度销售趋势", fontsize=15)
plt.xlabel("月份")
plt.ylabel("销售额（元）")
plt.grid(True, alpha=0.3)
plt.xticks(rotation=30)
for x, y in zip(by_month.index, by_month.values):
    plt.annotate(f"{y:,}", (x, y), textcoords="offset points", xytext=(0, 10),
                 ha="center", fontsize=10)
plt.tight_layout()
plt.savefig(report_dir / "chart_trend.png", dpi=120, bbox_inches="tight")
plt.close()

# 4.2 销售员排名
plt.figure(figsize=(8, 5))
bars = plt.bar(by_person.index, by_person.values, color=COLORS)
plt.title("销售员总业绩排名", fontsize=15)
plt.xlabel("销售员")
plt.ylabel("总销售额（元）")
for bar, v in zip(bars, by_person.values):
    plt.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1000,
             f"{v:,}", ha="center", fontsize=10)
plt.tight_layout()
plt.savefig(report_dir / "chart_person.png", dpi=120, bbox_inches="tight")
plt.close()

# 4.3 业绩占比
plt.figure(figsize=(8, 6))
plt.pie(by_person.values, labels=by_person.index, autopct="%1.1f%%",
        colors=COLORS, startangle=90, textprops={"fontsize": 11})
plt.title("销售员业绩贡献占比", fontsize=15)
plt.axis("equal")
plt.tight_layout()
plt.savefig(report_dir / "chart_pie.png", dpi=120, bbox_inches="tight")
plt.close()

# 4.4 每人月度走势对比
plt.figure(figsize=(10, 5))
for person, color in zip(by_person.index, COLORS):
    s = df[df["销售员"] == person].groupby("月份")["销售额"].sum()
    plt.plot(s.index, s.values, marker="o", label=person,
             linewidth=2, color=color, markersize=6)
plt.title("每人月度销售趋势对比", fontsize=15)
plt.xlabel("月份")
plt.ylabel("销售额（元）")
plt.legend(title="销售员")
plt.grid(True, alpha=0.3)
plt.xticks(rotation=30)
plt.tight_layout()
plt.savefig(report_dir / "chart_compare.png", dpi=120, bbox_inches="tight")
plt.close()

print("4 张图表已保存到 report/\n")

# ===== 5. 生成 Excel 数据报告 =====
wb = Workbook()

# Sheet 1: 数据概览
ws1 = wb.active
ws1.title = "数据概览"
ws1.append(["指标", "数值"])
ws1.append(["总销售额(元)", f"{total:,.0f}"])
ws1.append(["成交笔数", n_deals])
ws1.append(["平均客单价(元)", f"{avg_per_deal:,.0f}"])
ws1.append(["销售员人数", len(by_person)])
ws1.append(["月份数", len(by_month)])
ws1.append(["销售冠军", f"{top}"])
ws1.append(["冠军业绩(元)", f"{top_amount:,.0f}"])
ws1.append(["冠军占比", f"{top_pct:.1f}%"])
ws1.append(["6月较1月增长", f"{growth:.1f}%"])

# Sheet 2: 按销售员
ws2 = wb.create_sheet("按销售员")
ws2.append(["销售员", "总销售额(元)", "平均单笔(元)", "成交笔数", "业绩占比"])
for person in by_person.index:
    sub = df[df["销售员"] == person]
    ws2.append([
        person,
        f"{sub['销售额'].sum():,.0f}",
        f"{sub['销售额'].mean():,.0f}",
        len(sub),
        f"{sub['销售额'].sum() / total * 100:.1f}%",
    ])

# Sheet 3: 按月份
ws3 = wb.create_sheet("按月份")
ws3.append(["月份", "总销售额(元)", "环比增长"])
prev = 0
for month in by_month.index:
    curr = by_month[month]
    growth_pct = (curr - prev) / prev * 100 if prev > 0 else None
    ws3.append([
        str(month),
        f"{curr:,.0f}",
        f"{growth_pct:.1f}%" if growth_pct is not None else "—",
    ])
    prev = curr

# Sheet 4: 文字总结
ws4 = wb.create_sheet("分析报告")
ws4.append([f"销售分析报告 {date_str}"])
ws4.append([])
ws4.append(["【核心发现】"])
ws4.append([f"1. 上半年共完成 {n_deals} 笔销售，累计销售额 {total:,.0f} 元。"])
ws4.append([f"2. 销售冠军：{top}，个人贡献 {top_amount:,.0f} 元，占总业绩 {top_pct:.1f}%。"])
ws4.append([f"3. 销售趋势：6 月销售额 {by_month.iloc[-1]:,.0f} 元，较 1 月 {by_month.iloc[0]:,.0f} 元增长 {growth:.1f}%。"])
ws4.append([f"4. 平均客单价 {avg_per_deal:,.0f} 元，建议关注 {avg_per_deal < 10000 and '低客单' or '高客单'}客户群。"])
ws4.append([])
ws4.append(["【销售员排名】"])
for i, (person, amount) in enumerate(by_person.items(), 1):
    ws4.append([f"  {i}. {person}：{amount:,.0f} 元"])

# Sheet 5: 图表（嵌入 4 张 PNG）
ws5 = wb.create_sheet("图表")
chart_files = ["chart_trend.png", "chart_person.png", "chart_pie.png", "chart_compare.png"]
for i, fname in enumerate(chart_files):
    img = XLImage(str(report_dir / fname))
    img.width = 600
    img.height = 350
    ws5.add_image(img, f"A{1 + i * 22}")          # 每张图之间留 22 行

report_path = report_dir / f"销售分析报告_{date_str}.xlsx"
wb.save(report_path)
print(f"Excel 报告已生成：{report_path}")
print(f"\n🎉 报告生成完成！请用 WPS 打开查看")
