# read_excel.py —— 读取 sales.xlsx 的 4 种方式
from openpyxl import load_workbook

wb = load_workbook("sales.xlsx")
ws = wb.active

print(f"当前工作表：{ws.title}")
print(f"最大行：{ws.max_row}，最大列：{ws.max_column}")
print("-" * 30)

print("方式1：按单元格地址读取")
print("A1 =", ws["A1"].value)
print("B2 =", ws["B2"].value)
print()

print("方式2：按行号列号读取（都从 1 开始）")
print("第1行第1列 =", ws.cell(row=1, column=1).value)
print("第2行第2列 =", ws.cell(row=2, column=2).value)
print()

print("方式3：遍历所有行（values_only=True 直接拿值）")
for row in ws.iter_rows(values_only=True):
    print(row)
print()

print("方式4：遍历某一列")
for cell in ws["B"]:
    print("B列：", cell.value)
