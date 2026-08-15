# make_report.py —— 读 sales.xlsx，生成"月度汇总.xlsx"
# 这是第 2 节课的演示：读 + 处理 + 写 完整流程

from openpyxl import load_workbook, Workbook

# 1. 读取原数据
src = load_workbook("sales.xlsx")
src_ws = src.active

# 2. 创建新工作簿 + 新表
dst = Workbook()
ws = dst.active
ws.title = "月度汇总"

# 3. 写表头
ws.append(["月份", "销售额", "利润", "利润率"])

# 4. 逐行读取原表，写入新表 + 同时算利润率
for row in src_ws.iter_rows(min_row=2, values_only=True):  # min_row=2 跳过表头
    month, sales, profit = row
    rate = profit / sales
    ws.append([month, sales, profit, rate])

# 5. 最后追加"合计"行
sales_list = [row[1] for row in src_ws.iter_rows(min_row=2, values_only=True)]
profit_list = [row[2] for row in src_ws.iter_rows(min_row=2, values_only=True)]
total_sales = sum(sales_list)
total_profit = sum(profit_list)
ws.append(["合计", total_sales, total_profit, total_profit / total_sales])

# 6. 保存
dst.save("月度汇总.xlsx")
print("月度汇总.xlsx 已生成 ✅")

# 7. 再读出来看一眼，确认内容正确
print("\n--- 新文件内容 ---")
check = load_workbook("月度汇总.xlsx")
for row in check.active.iter_rows(values_only=True):
    print(row)