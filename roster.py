from openpyxl import Workbook, load_workbook

# 1. 创建 + 写表头
wb = Workbook()
ws = wb.active
ws.title = "员工花名册"
ws.append(["姓名", "部门", "工资", "入职日期"])

# 2. 写数据（列表套元组，一个员工一行）
employees = [
    ("张三", "技术部",  8000, "2024-03-15"),
    ("李四", "市场部",  6500, "2025-01-08"),
    ("王五", "财务部",  7000, "2023-11-20"),
    ("赵六", "技术部",  8500, "2024-07-01"),
    ("钱七", "人事部",  6000, "2025-05-12"),
]

for emp in employees:
    ws.append(emp)

wb.save("员工花名册.xlsx")
print("员工花名册.xlsx 已生成 ✅")

# 3. 立刻读回来看一眼（自检）
print("\n--- 内容预览 ---")
check = load_workbook("员工花名册.xlsx")
for row in check.active.iter_rows(values_only=True):
    print(row)
