"""月度销售报表生成器 - 综合实战项目"""
from pathlib import Path
from datetime import datetime
from collections import defaultdict
from shutil import copy2

from openpyxl import Workbook, load_workbook

# ============== 配置区（用的人只改这里就行）======================
SOURCE_DIR = Path("monthly_sales")           # 月度数据文件所在文件夹
OUTPUT_FILE = Path("月度销售汇总.xlsx")      # 输出的总表
BACKUP_DIR = Path("backup")                  # 备份目录
# ============================================================


def scan_source_files(source_dir):
    """扫描源文件夹，找出所有符合格式的 Excel 文件"""
    if not source_dir.exists():
        print(f"❌ 源文件夹不存在：{source_dir}")
        return []

    files = sorted(source_dir.glob("sales_*.xlsx"))
    print(f"📁 扫描到 {len(files)} 个月度数据文件")
    return files


def backup_files(files, backup_dir):
    """把原始文件备份到 backup/YYYY-MM-DD/"""
    today = datetime.now().strftime("%Y-%m-%d")
    dest = backup_dir / today
    dest.mkdir(parents=True, exist_ok=True)

    for f in files:
        copy2(f, dest / f.name)
    print(f"💾 已备份到 {dest}/  ({len(files)} 个文件)")


def merge_all_data(files):
    """合并所有 Excel 的明细数据"""
    all_rows = []
    for f in files:
        try:
            wb = load_workbook(f)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] is None:
                    continue
                all_rows.append(row)
            wb.close()
        except Exception as e:
            print(f"⚠️  跳过 {f.name}：{e}")
    print(f"📊 共合并 {len(all_rows)} 条销售记录")
    return all_rows


def summarize_by_person(all_rows):
    """按销售员汇总"""
    totals = defaultdict(int)
    for _, person, amount in all_rows:
        totals[person] += amount

    rows = [("销售员", "总销售额")]
    for person, total in sorted(totals.items(), key=lambda x: -x[1]):
        rows.append((person, total))
    return rows, totals


def summarize_by_month(all_rows):
    """按月份汇总"""
    totals = defaultdict(int)
    for month, _, amount in all_rows:
        totals[month] += amount

    rows = [("月份", "总销售额")]
    for month, total in sorted(totals.items()):
        rows.append((month, total))
    return rows, totals


def write_report(all_rows, by_person, by_month, output_file):
    """生成 3 张工作表的报表"""
    wb = Workbook()

    # 第 1 张：所有明细
    ws1 = wb.active
    ws1.title = "明细"
    ws1.append(("月份", "销售员", "销售额"))
    for row in all_rows:
        ws1.append(row)

    # 第 2 张：按销售员汇总
    ws2 = wb.create_sheet("按销售员")
    for row in by_person:
        ws2.append(row)

    # 第 3 张：按月份汇总
    ws3 = wb.create_sheet("按月份")
    for row in by_month:
        ws3.append(row)

    wb.save(output_file)
    print(f"✅ 报表已生成：{output_file}")


def main():
    print("=" * 50)
    print("🚀 月度销售报表生成器 启动")
    print("=" * 50)

    files = scan_source_files(SOURCE_DIR)
    if not files:
        return

    backup_files(files, BACKUP_DIR)

    all_rows = merge_all_data(files)
    if not all_rows:
        print("❌ 没有可用数据，程序退出")
        return

    by_person, _ = summarize_by_person(all_rows)
    by_month, _  = summarize_by_month(all_rows)

    write_report(all_rows, by_person, by_month, OUTPUT_FILE)

    # 终端简报
    print("\n📈 终端简报：")
    for person, total in by_person[1:]:
        print(f"  {person}: ¥{total}")
    print("=" * 50)


if __name__ == "__main__":
    main()